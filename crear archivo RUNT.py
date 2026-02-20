import os
import shutil
from openpyxl import load_workbook

def modificar_archivo_excel(archivo_destino):
    # Cargar el archivo Excel
    try:
        wb = load_workbook(archivo_destino)
        sheet = wb.active
        
        # Modificar el nombre de la columna G a "RUNT"
        sheet['G1'] = 'RUNT'  # Suponiendo que la primera fila es el encabezado
        
        # Asignar el valor "SIN RUNT" a todas las celdas en la columna G (excepto el encabezado)
        for row in range(2, sheet.max_row + 1):  # Empezamos desde la fila 2 para no modificar el encabezado
            sheet[f'G{row}'] = 'SIN RUNT'

        # Guardar los cambios en el archivo
        wb.save(archivo_destino)
        print(f"Archivo modificado: {archivo_destino}")
    except Exception as e:
        print(f"Error al modificar el archivo {archivo_destino}: {e}")

def copiar_archivos_excel(ruta_base):
    # Recorre todas las carpetas en la ruta base
    for root, dirs, files in os.walk(ruta_base):
        # Filtramos solo los archivos Excel (.xlsx o .xls)
        archivos_excel = [f for f in files if f.lower().endswith(('.xlsx', '.xls'))]
        
        # Si encontramos un archivo con "RUNT", pasamos a la siguiente carpeta
        for archivo in archivos_excel:
            if "RUNT" in archivo:
                # Si el archivo contiene "RUNT", salimos del ciclo de archivos y pasamos a la siguiente carpeta
                print(f"Archivo {archivo} contiene 'RUNT', se ignora esta carpeta.")
                break  # Salta a la siguiente carpeta

        else:
            # Si no se encontró ningún archivo con "RUNT", procesamos los archivos sin "_"
            for archivo in archivos_excel:
                # Verifica si el archivo no contiene el carácter "_"
                if "_" not in archivo:
                    # Construir la ruta del archivo original
                    archivo_original = os.path.join(root, archivo)
                    
                    # Aseguramos que se cambie el nombre correctamente dependiendo de la extensión
                    if archivo.lower().endswith(".xlsx"):
                        nuevo_nombre = archivo.replace(".xlsx", "_RUNT.xlsx")
                    elif archivo.lower().endswith(".xls"):
                        nuevo_nombre = archivo.replace(".xls", "_RUNT.xls")
                    archivo_destino = os.path.join(root, nuevo_nombre)

                    # Intentamos copiar el archivo
                    try:
                        shutil.copy(archivo_original, archivo_destino)
                        print(f"Archivo copiado: {archivo_original} -> {archivo_destino}")
                        
                        # Modificar el archivo Excel copiado
                        modificar_archivo_excel(archivo_destino)
                        
                    except Exception as e:
                        print(f"Error al copiar el archivo {archivo_original}: {e}")

# Ruta base donde se encuentran las carpetas
ruta_base = r"C:\Users\josea.gonzalez\Documents\CONTRALORIA\UNCOPI\UNCOPI 2025\209\19 de mayo de 2025_209"  # Aquí debes poner la ruta correcta

# Llamamos a la función
copiar_archivos_excel(ruta_base)
